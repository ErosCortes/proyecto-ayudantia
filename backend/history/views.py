from django.http import HttpResponse
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from users.permissions import EsAlumno, EsProfesor, EsAdmin
from rest_framework.permissions import IsAuthenticated
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import AssistantHistory
from .serializers import AssistantHistorySerializer


class AssistantHistoryViewSet(viewsets.ModelViewSet):
    queryset = AssistantHistory.objects.all()
    serializer_class = AssistantHistorySerializer
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        if self.action in ['my_history', 'by_semester']:
            return [EsAlumno()]
        if self.action == 'by_student':
            return [EsProfesor()]
        if self.action in ['all_history', 'export_excel']:
            return [EsAdmin()]
        return [IsAuthenticated()]

    @action(detail=False, methods=['get'])
    def my_history(self, request):
        """Historial del alumno actual (solo para el propio alumno)"""
        history = AssistantHistory.objects.filter(
            id_alumno=request.user
        ).select_related('id_curso__course')
        serializer = self.get_serializer(history, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def by_semester(self, request):
        """Historial del alumno actual filtrado por semestre"""
        semestre = request.query_params.get('semestre')
        history = AssistantHistory.objects.filter(
            id_alumno=request.user,
            semestre=semestre
        ).select_related('id_curso__course')
        serializer = self.get_serializer(history, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def by_student(self, request):
        """
        Historial de un alumno específico — accesible para profesores.
        Uso: GET /api/history/by_student/?student_id=<id>
        """
        student_id = request.query_params.get('student_id')
        if not student_id:
            return Response({'error': 'Se requiere el parámetro student_id'}, status=400)
        history = AssistantHistory.objects.filter(
            id_alumno_id=student_id
        ).select_related('id_alumno', 'id_curso__course')
        serializer = self.get_serializer(history, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def all_history(self, request):
        """
        Historial completo con filtros — solo administradores.
        Filtros opcionales: ?alumno=<nombre_o_rut>&semestre=<semestre>&estado=<estado>&curso=<nombre>
        """
        qs = AssistantHistory.objects.select_related(
            'id_alumno', 'id_curso__course'
        ).order_by('-semestre', 'id_alumno__nombre_completo')
        alumno = request.query_params.get('alumno', '').strip()
        semestre = request.query_params.get('semestre', '').strip()
        estado = request.query_params.get('estado', '').strip()
        curso = request.query_params.get('curso', '').strip()
        if alumno:
            qs = qs.filter(
                id_alumno__nombre_completo__icontains=alumno
            ) | qs.filter(
                id_alumno__rut__icontains=alumno
            )
        if semestre:
            qs = qs.filter(semestre=semestre)
        if estado:
            qs = qs.filter(estado_final=estado)
        if curso:
            qs = qs.filter(id_curso__course__nombre__icontains=curso)
        serializer = self.get_serializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """
        Exporta a Excel el historial de ayudantías, pensado para el proceso
        de pago. No incluye datos bancarios ni sensibles, solo lo necesario
        para identificar al ayudante y la sección.
        Filtro opcional: ?estado=COMPLETADA|INCOMPLETA|CANCELADA|TODOS
        (por defecto exporta COMPLETADA)
        """
        estado_filtro = request.query_params.get('estado', 'COMPLETADA')

        historial = AssistantHistory.objects.select_related(
            'id_alumno', 'id_curso', 'id_curso__course'
        )
        if estado_filtro and estado_filtro != 'TODOS':
            historial = historial.filter(estado_final=estado_filtro)

        wb = Workbook()
        ws = wb.active
        ws.title = "Ayudantías"

        headers = [
            "Nombre",
            "RUT",
            "Correo institucional",
            "Código curso",
            "Nombre curso",
            "NRC",
            "Semestre",
            "Año",
            "Estado",
            "Fecha registro",
        ]
        ws.append(headers)

        header_fill = PatternFill(start_color="003057", end_color="003057", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_num, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for h in historial:
            alumno = h.id_alumno
            section = h.id_curso
            course = section.course if section else None

            ws.append([
                alumno.nombre_completo,
                alumno.rut,
                alumno.correo_institucional,
                course.codigo_curso if course else "",
                course.nombre if course else "",
                section.nrc if section else "",
                h.semestre,
                getattr(section, "year", ""),
                h.get_estado_final_display(),
                h.created_at.strftime("%d-%m-%Y") if h.created_at else "",
            ])

        for col_num, header in enumerate(headers, start=1):
            max_length = len(header)
            for row in ws.iter_rows(min_col=col_num, max_col=col_num, min_row=2):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_num)].width = max_length + 3

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = "ayudantias_pago.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response